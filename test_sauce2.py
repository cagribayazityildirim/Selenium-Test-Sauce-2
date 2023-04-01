from typing import KeysView
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait #bekleme işlemlerini ele yapan bir yapı 
from selenium.webdriver.support import expected_conditions as ec #hangi şarta göre bekleyeceğimizi söylemek için
import pytest
from pathlib import Path
from datetime import date
import globalConstants
import openpyxl

class Test_Sauce2:
    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
        
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.waitForElementVisible((By.ID,globalConstants.userNameInput))
        self.waitForElementVisible((By.ID,globalConstants.passwordInput))   
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)  

    def teardown_method(self):
        self.driver.quit()

    def success_login(self):
        usernameInput = self.driver.find_element(By.ID,globalConstants.userNameInput)
        usernameInput.click()
        usernameInput.send_keys("standard_user")
        passwordInput = self.driver.find_element(By.ID,globalConstants.passwordInput)
        passwordInput.click()
        passwordInput.send_keys("secret_sauce")
        loginBTN = self.driver.find_element(By.ID,globalConstants.loginButton)
        loginBTN.click()
    
    @pytest.mark.parametrize("username,password",[("standard_user","secret_sauce")])
    def test_success_login(self,username,password):
        usernameInput = self.driver.find_element(By.ID,globalConstants.userNameInput)
        usernameInput.click()
        usernameInput.send_keys(username)
        passwordInput = self.driver.find_element(By.ID,globalConstants.passwordInput)
        passwordInput.click()
        passwordInput.send_keys(password)
        loginBTN = self.driver.find_element(By.ID,globalConstants.loginButton)
        loginBTN.click(),
        self.driver.save_screenshot(f"{self.folderPath}/test-success-login.png")
        assert self.driver.current_url == globalConstants.URL2
    
    def getData():
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["Sheet1"]

        totalRows = selectedSheet.max_row
        data = []
        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
            
        return data


    @pytest.mark.parametrize("username,password",getData())
    def test_all_user_login(self,username,password):
        usernameInput = self.driver.find_element(By.ID,globalConstants.userNameInput)
        usernameInput.click()
        usernameInput.send_keys(username)
        passwordInput = self.driver.find_element(By.ID,globalConstants.passwordInput)
        passwordInput.click()
        passwordInput.send_keys(password)
        loginBTN = self.driver.find_element(By.ID,globalConstants.loginButton)
        loginBTN.click(),
        self.driver.save_screenshot(f"{self.folderPath}/test-all-user-login.png")
        assert self.driver.current_url == globalConstants.URL2


    def test_product_number(self):
        self.success_login()
        self.waitForElementVisible((By.ID,globalConstants.menuBtn))
        listSauce = self.driver.find_elements(By.CLASS_NAME,"inventory_item")
        self.driver.save_screenshot(f"{self.folderPath}/test-product-number.png")
        assert len(listSauce) == 6

    def test_success_logout(self):
        self.success_login()
        self.waitForElementVisible((By.ID,globalConstants.menuBtn))
        menuBtn = self.driver.find_element(By.ID,globalConstants.menuBtn).click()
        self.waitForElementVisible((By.ID,globalConstants.logoutBtn))
        logoutBtn = self.driver.find_element(By.ID,globalConstants.logoutBtn).click()
        self.driver.save_screenshot(f"{self.folderPath}/test-success-logout.png")
        assert self.driver.current_url == globalConstants.URL
       
    def test_labsbackpack_add_cart(self):
        self.success_login()
        self.waitForElementVisible((By.ID,globalConstants.menuBtn))
        addcartBtnbackpack = self.driver.find_element(By.ID,globalConstants.addcartBtnbackpack)
        addcartBtnbackpack.click()
        shoppingCart = self.driver.find_element(By.CLASS_NAME,globalConstants.shoppingCart)
        shoppingCart.click()
        self.waitForElementVisible((By.CLASS_NAME,globalConstants.addcartBtnbackpackClass))
        labsBackpackName = self.driver.find_element(By.CLASS_NAME,globalConstants.addcartBtnbackpackClass)
        self.driver.save_screenshot(f"{self.folderPath}/test-labBackpack-add-cart.png")
        assert labsBackpackName.text == globalConstants.labsBackpackName

    def test_add_remove(self):
        self.success_login()
        self.waitForElementVisible((By.ID,globalConstants.menuBtn))
        addcartBtnbackpack = self.driver.find_element(By.ID,globalConstants.addcartBtnbackpack)
        addcartBtnbackpack.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-add1-remove.png")
        removeBtnbackpack = self.driver.find_element(By.ID,globalConstants.removeBtnbackpack)
        removeBtnbackpack.click()
        addcartBtnbackpack = self.driver.find_element(By.ID,globalConstants.addcartBtnbackpack)
        self.driver.save_screenshot(f"{self.folderPath}/test-add-remove.png")
        assert addcartBtnbackpack.text == globalConstants.addcartBtntext

    def test_low_high(self):
        self.success_login()
        self.waitForElementVisible((By.ID,globalConstants.menuBtn))
        sortBtn = self.driver.find_element(By.CLASS_NAME,globalConstants.sortBtn)
        sortBtn.click()
        low_highBtn = self.driver.find_element(By.XPATH,globalConstants.low_highBtn)
        low_highBtn.click()
        self.waitForElementVisible((By.CLASS_NAME,globalConstants.labsOnesieclassname))
        productFirst = self.driver.find_element(By.CLASS_NAME,globalConstants.labsOnesieclassname)
        result = productFirst.text
        self.driver.save_screenshot(f"{self.folderPath}/test-low-high.png")
        assert result == globalConstants.labsOnesietext

    def test_high_low(self):
        self.success_login()
        self.waitForElementVisible((By.ID,globalConstants.menuBtn))
        sortBtn = self.driver.find_element(By.CLASS_NAME,globalConstants.sortBtn)
        sortBtn.click()
        high_lowBtn = self.driver.find_element(By.XPATH,globalConstants.high_lowBtn)
        high_lowBtn.click()
        self.waitForElementVisible((By.CLASS_NAME,globalConstants.labsfliceJacket))
        productFirst = self.driver.find_element(By.CLASS_NAME,globalConstants.labsfliceJacket)
        result = productFirst.text
        self.driver.save_screenshot(f"{self.folderPath}/test-high-low.png")
        assert result == globalConstants.labsfliceJacketName
